# Copyright (c) 2023, Nirali and contributors
# For license information, please see license.txt

import frappe,json
from frappe import _
from frappe.utils import get_link_to_form
from frappe.model.document import Document
from frappe.model.mapper import get_mapped_doc
from erpnext.controllers.item_variant import (
	ItemVariantExistsError,
	copy_attributes_to_variant,
	get_variant,
	make_variant_item_code,
	validate_item_variant_attributes,
)
from frappe.utils import (
	cint,
	cstr,
	flt,
	formatdate,
	get_link_to_form,
	getdate,
	now_datetime,
	nowtime,
	strip,
	strip_html,
)
import frappe
from frappe.utils import now_datetime, get_datetime
from frappe.utils import get_link_to_form
from frappe import _
from datetime import datetime, time, timedelta
from frappe.utils.file_manager import save_file
import openpyxl
from io import BytesIO
import requests
import os

class OrderForm(Document):
	
	def on_submit(self):
		create_cad_orders(self)
		if self.supplier:
			create_po(self)


	def on_update_after_submit(self):
		if self.updated_delivery_date:
			order_names = frappe.get_all(
				"Order",
				filters={"cad_order_form": self.name},
				pluck="name"
			)

			for order_name in order_names:
				frappe.db.set_value("Order", order_name, "updated_delivery_date", self.updated_delivery_date)

	def on_cancel(self):
		order_names = frappe.db.get_list(
			"Order",
			filters={"cad_order_form": self.name}, 
			pluck ="name")
		if order_names:
			frappe.db.set_value(
				"Order",
				{"name"["in",order_names]},
				"workflow_state",
				"Cancelled"
			)
			timesheet_names = frappe.db.get_list(
				"Timesheet",
				filters={"order": ["in", order_names]},
				pluck="name"
			)
			if timesheet_names:
				frappe.db.set_value(
					"Timesheet",
					{"name": ["in", timesheet_names]},
					"workflow_state",
					"Cancelled"
				)
		frappe.db.set_value(
			"Order Form",
			self.name,
			"workflow_state",
			"Cancelled"
		)

	def validate(self):
		self.validate_category_subcategory()
		self.validate_field_value()
		validate_design_id(self)
		validate_item_variant(self)
		validate_is_mannual(self)
		set_data(self)
		for i in self.order_details:	
			if i.metal_type == "Silver":
				i.metal_colour = "White"
				i.metal_touch = "20KT"
				i.setting_type = "Open"
				i.diamond_type = "AD"

	def validate_category_subcategory(self):
		for row in self.get("order_details"):
			if row.subcategory:
				parent = frappe.db.get_value("Attribute Value", row.subcategory, "parent_attribute_value")
				if row.category != parent:
					frappe.throw(_(f"Category & Sub Category mismatched in row #{row.idx}"))
	
	def validate_field_value(self):
		field_attribute_map = {
			"design_type": "Design Type",
			"diamond_quality": "Diamond Quality",
			"setting_type": "Setting Type",
			"sub_setting_type1": "Sub Setting Type1",
			"sub_setting_type2": "Sub Setting Type2",
			"metal_type": "Metal Type",
			"metal_touch": "Metal Touch",
			"metal_colour": "Metal Colour",
			"diamond_type": "Diamond Type",
			"sizer_type": "Sizer Type",
			"stone_changeable": "Stone Changeable",
			"feature": "Feature",
			"rhodium": "Rhodium",
			"enamal": "Enamal",
			"gemstone_type": "Gemstone Type",
			"gemstone_quality": "Gemstone Quality",
			"mod_reason": "Mod Reason",
			"finding_category": "Finding Category",
			"finding_subcategory": "Finding Sub-Category",
			"finding_size": "Finding Size",
			"metal_target_from_range": "Metal Target Range",
			"diamond_target_from_range": "Diamond Target Range",
			"detachable": "Detachable",
			"lock_type": "Lock Type",
			"capganthan": "Cap/Ganthan",
			"charm": "Charm",
			"back_chain": "Back Chain",
			"back_belt": "Back Belt",
			"black_bead": "Black Bead",
			"two_in_one": "2 in 1",
			"chain_type": "Chain Type",
			"nakshi_from": "Nakshi From",
		}
		all_attributes = frappe.get_all(
			"Item Attribute Value",
			fields=["parent", "attribute_value"]
		)

		attribute_lookup = {}
		for attr in all_attributes:
			if attr.parent not in attribute_lookup:
				attribute_lookup[attr.parent] = []
			attribute_lookup[attr.parent].append(attr.attribute_value)

		row_no = 0
		for row in self.get("order_details"):
			row_no += 1

			for field in field_attribute_map:
				value = row.get(field)
				if value:
					attribute_name = field_attribute_map[field]

					if value not in attribute_lookup.get(attribute_name, []):
						frappe.throw(
							"Row {0}: {1} is not Correct".format(
								row_no, attribute_name
							)
						)

def create_cad_orders(self):
    
    if self.docstatus == 0 or self.workflow_state in ["Draft","Send For Approval", "Cancelled"]:
        frappe.msgprint(_("Order creation skipped because document is in Draft or Cancelled state."))
        return

    doclist = []

    order_criteria = frappe.get_single("Order Criteria")
    criteria_rows = order_criteria.get("order")
    enabled_criteria = next((row for row in criteria_rows if not row.disable), None)

    if not enabled_criteria:
        frappe.throw("No enabled Order Criteria found.")

    cad_days = int(enabled_criteria.cad_approval_day or 0)

    cad_time_raw = enabled_criteria.cad_submission_time
    if isinstance(cad_time_raw, time):
        cad_time = cad_time_raw
    elif isinstance(cad_time_raw, timedelta):
        cad_time = (datetime.min + cad_time_raw).time()
    elif isinstance(cad_time_raw, str):
        try:
            h, m, s = [int(x) for x in cad_time_raw.strip().split(".")]
            cad_time = time(h, m, s)
        except:
            frappe.throw("Invalid CAD Submission Time format.")
    else:
        cad_time = time(0, 0, 0)

    ibm_time_raw = enabled_criteria.cad_appoval_timefrom_ibm_team
    if isinstance(ibm_time_raw, time):
        ibm_timedelta = timedelta(hours=ibm_time_raw.hour, minutes=ibm_time_raw.minute, seconds=ibm_time_raw.second)
    elif isinstance(ibm_time_raw, timedelta):
        ibm_timedelta = ibm_time_raw
    elif isinstance(ibm_time_raw, str):
        try:
            h, m, s = [int(x) for x in ibm_time_raw.strip().split(".")]
            ibm_timedelta = timedelta(hours=h, minutes=m, seconds=s)
        except:
            frappe.throw("Invalid IBM Approval Time format.")
    else:
        ibm_timedelta = timedelta()

    for row in self.order_details:
        docname = make_cad_order(row.name, parent_doc=self)

        if row.pre_order_form_details:
            frappe.db.set_value("Pre Order Form Details", row.pre_order_form_details, "order_form_id", self.name)

        order_datetime = now_datetime()
        frappe.db.set_value("Order", docname, "order_date", order_datetime)

        if self.delivery_date:
            frappe.db.set_value("Order", docname, "delivery_date", self.delivery_date)

        cad_delivery_datetime = datetime.combine(order_datetime.date() + timedelta(days=cad_days), cad_time)
        ibm_delivery_datetime = cad_delivery_datetime + ibm_timedelta

        frappe.db.set_value("Order", docname, "cad_delivery_date", cad_delivery_datetime)
        frappe.db.set_value("Order", docname, "ibm_delivery_date", ibm_delivery_datetime)

        doclist.append(get_link_to_form("Order", docname))

    if doclist:
        msg = _("The following {0} were created: {1}").format(
            frappe.bold(_("Orders")), "<br>" + ", ".join(doclist)
        )
        frappe.msgprint(msg)

def make_cad_order(source_name, target_doc=None, parent_doc = None):
	def set_missing_values(source, target):
		target.cad_order_form_detail = source.name
		target.cad_order_form = source.parent
		target.index = source.idx
	source_doc = frappe.get_doc('Order Form Detail',source_name)
	design_type = source_doc.design_type
	item_type = source_doc.item_type
	is_repairing = source_doc.is_repairing
	is_finding_order = source_doc.is_finding_order

	if design_type == 'Mod - Old Stylebio & Tag No':
		if is_repairing == 1:
			bom_or_cad = source_doc.bom_or_cad
			item_type = source_doc.item_type
		else:
			bom_or_cad = 'Check'
	elif design_type == 'Sketch Design':
		item_type = "Only Variant"
		bom_or_cad = 'CAD'
	elif design_type == 'As Per Design Type':
		item_type = "No Variant No Suffix"
		bom_or_cad = 'New BOM'
	elif is_finding_order:
		item_type = "No Variant No Suffix"
		bom_or_cad = 'New BOM'
	else:
		item_type = 'Template and Variant'
		bom_or_cad = 'CAD'

	doc = get_mapped_doc(
		"Order Form Detail",
		source_name,
		{
			"Order Form Detail": {
				"doctype": "Order" 
			}
		},target_doc, set_missing_values
	)

	if parent_doc:
		for entity in parent_doc.get("service_type",[]):
			doc.append("service_type", {"service_type1": entity.service_type1})
		
		doc.update({
			"parcel_place":parent_doc.parcel_place,
			"company":parent_doc.company,
			"form_remarks":parent_doc.remarks,
			"india":parent_doc.india,
			"usa":parent_doc.usa,
			"india_states":parent_doc.india_states,
			"item_type":item_type,
			"bom_or_cad":bom_or_cad
		})
	if design_type in ['New Design','Sketch Design']:
		doc.workflow_type = 'CAD'
	doc.save()
	if design_type == 'As Per Design Type' and item_type == "No Variant No Suffix" and bom_or_cad == 'New BOM':
		doc.submit()
		frappe.db.set_value("Order",doc.name,"workflow_state","Approved")
	return doc.name
 
@frappe.whitelist()
def get_sketch_details(design_id):

    final_data = {}
    data = frappe.db.sql(
        """
        SELECT
            i.item_category,
            i.item_subcategory,
            i.setting_type,
            i.approx_gold        AS metal_target,
            i.approx_diamond    AS diamond_target,

            so.sub_setting_type1,
            so.sub_setting_type2,
            so.qty,
            so.metal_type,
            so.metal_touch,
            so.metal_colour,
            so.product_size,
            so.sizer_type,
            so.length,
            so.width,
            so.height

        FROM `tabItem` i
        LEFT JOIN `tabSketch Order` so
            ON so.name = i.custom_sketch_order_id
        WHERE i.name = %s
        """,
        (design_id,),
        as_dict=True
    )

    if data:
        final_data.update(data[0])
    db_data = frappe.db.get_all(
        "Item Variant Attribute",
        filters={"parent": design_id},
        fields=["attribute", "attribute_value"]
    )

    for i in db_data:
        if not i.attribute_value:
            continue

        final_data[i.attribute.lower().replace(" ", "_")] = i.attribute_value

    return final_data

@frappe.whitelist()
def get_customer_orderType(customer_code):
	order_type = frappe.db.sql(
		f""" select order_type from `tabOrder Type` where parent= '{customer_code}' """, as_dict=1
	)

	return order_type

@frappe.whitelist()
def get_customer_order_form(source_name, target_doc=None):
	if isinstance(target_doc, str):
		target_doc = json.loads(target_doc)
	target_doc = frappe.new_doc("Order Form") if not target_doc else frappe.get_doc(target_doc)

	if source_name:
		customer_order_form = frappe.db.sql(f"""SELECT * FROM `tabCustomer Order Form Detail` 
							WHERE parent = '{source_name}' AND docstatus = 1""", as_dict=1)
	if not customer_order_form:
		frappe.msgprint(_("Please submit the Customer Order Form"))
		return target_doc

	for i in customer_order_form:
		item, order_id, item_bom = i.get("design_code"), i.get("order_id"), i.get("design_code_bom")
		order_data = frappe.db.sql(f"SELECT * FROM `tabOrder` WHERE name = '{order_id}'", as_dict=1)
		
		customer_design_code = frappe.db.sql(f"SELECT * FROM `tabBOM` WHERE item = '{item}' AND name = '{i.get('design_code_bom')}'", as_dict=1)
		item_serial = frappe.db.get_value("Serial No", {'item_code': item}, 'name')
		
		data_source = order_data if order_data else customer_design_code
		
		product_code = ''
		if i.get("digit14_code"):
			product_code = i.get("digit14_code")
		elif i.get("digit18_code"):
			product_code = i.get("digit18_code")
		elif i.get("digit15_code"):
			product_code = i.get("digit15_code")
		elif i.get("sku_code"):
			product_code = i.get("sku_code")
		if data_source:
			for j in data_source:
				target_doc.append("order_details", {
					"delivery_date": target_doc.delivery_date,
					"design_by": j.get('design_by'),
					"design_type": j.get('design_type'),
					"qty": i.get('no_of_pcs'),
					"design_id": j.get("item", item),
					"bom": j.get("new_bom", i.get('design_code_bom')),
					"tag_no": item_serial or j.get('tag_no'),
					"diamond_quality": i.get("diamond_quality"),
					"customer_order_form": i.get("parent"),
					"category": i.get("category"),
					"subcategory": i.get("subcategory"),
					"setting_type": i.get("setting_type"),
					"product_code": product_code if product_code else '',
					"theme_code": i.get("theme_code"),
					"metal_type": i.get("metal_type"),
					"metal_touch": i.get("metal_touch"),
					"metal_colour": i.get("metal_colour"),
					"metal_target": i.get("metal_target"),
					"diamond_target": i.get("diamond_target"),
					"feature": i.get("feature"),
					"product_size": i.get("product_size"),
					"rhodium": i.get("rhodium"),
					"enamal": j.get("enamal"),
					"sub_setting_type1": j.get("sub_setting_type1"),
					"sub_setting_type2": j.get("sub_setting_type2"),
					"sizer_type": j.get("sizer_type"),
					"stone_changeable": j.get("stone_changeable"),
					"detachable": j.get("detachable"),
					"lock_type": j.get("lock_type"),
					"capganthan": j.get("capganthan"),
					"charm": j.get("charm"),
					"back_chain": j.get("back_chain"),
					"back_chain_size": j.get("back_chain_size"),
					"back_belt": j.get("back_belt"),
					"back_belt_length": j.get("back_belt_length"),
					"black_beed_line": j.get("black_beed_line"),
					"back_side_size": j.get("back_side_size"),
					"back_belt_patti": j.get("back_belt_patti"),
					"two_in_one": j.get("two_in_one"),
					"number_of_ant": j.get("number_of_ant"),
					"distance_between_kadi_to_mugappu": j.get("distance_between_kadi_to_mugappu"),
					"space_between_mugappu": j.get("space_between_mugappu"),
					"chain_type": j.get("chain_type"),
					"customer_chain": j.get("customer_chain"),
					"nakshi_weght": j.get("nakshi_weght"),
				})
		else: 
			frappe.throw(f"{item} has master bom {item_bom}")
	return target_doc


def validate_item_variant(self):
	for i in self.order_details:
		if i.design_type == "Sketch Design" and i.design_id:
			custom_sketch_order_id = frappe.db.get_value("Item", i.design_id, "custom_sketch_order_id")
			if custom_sketch_order_id:
				# Get all variants where variant_of = i.design_id
				variants = frappe.get_all("Item",
					filters={"variant_of": i.design_id},
					fields=["name"]
				)
				if variants:
					variant_names = ", ".join(item.name for item in variants)
					frappe.throw(f"""
						You already created a variant for this Design ID ({i.design_id}).<br><br>
						Items found: {variant_names}<br><br>
						You cannot create another variant using <b>Sketch Design</b>.<br>
						Please select <b>Design Type = 'Mod - Old Stylebio & Tag No'</b> and select the variant in Design ID.
					""")


def validate_design_id(self):
	for i in self.order_details:
		# If tagno exists, find all matching enabled Items by old_tag_no
		if i.tagno:
			matching_items = frappe.db.get_all(
				"Item",
				filters={"old_tag_no": i.tagno, "disabled": 0},
				fields=["name", "master_bom", "creation"],
				order_by="creation desc"
			)

			if matching_items:
				# Pick the latest enabled item
				item = matching_items[0]
				matched_design_id = item.name

				# Set design_id only if not manually overridden
				if not i.design_id or i.design_id == matched_design_id:
					i.design_id = matched_design_id
					if not i.bom:
						i.bom = item.master_bom

		# If design_id is set, fetch Item and set master_bom to bom
		if i.design_id:
			item_doc = frappe.get_doc("Item", i.design_id)
			if item_doc.master_bom and not i.bom:
				i.bom = item_doc.master_bom

		# Continue only if design_id and bom are now set
		if i.design_id and i.bom:
			is_manual_override = (i.design_type == "Mod - Old Stylebio & Tag No")

			# Skip if mod_reason is "Change In Metal Type"
			if i.mod_reason != "Change In Metal Type":
				bom_doc = frappe.get_doc("BOM", i.bom)

				# Set metal_type and metal_touch from metal_detail
				if bom_doc.metal_detail:
					if not is_manual_override or not i.metal_type:
						i.metal_type = bom_doc.metal_detail[0].metal_type or None
					if not is_manual_override or not i.metal_touch:
						i.metal_touch = bom_doc.metal_detail[0].metal_touch or None
				else:
					frappe.msgprint(f"No metal details found for BOM {i.bom}")

				# Set setting_type, category, subcategory
				if not is_manual_override or not i.setting_type:
					i.setting_type = bom_doc.setting_type or None
				if not is_manual_override or not i.category:
					i.category = bom_doc.item_category or None
				if not is_manual_override or not i.subcategory:
					i.subcategory = bom_doc.item_subcategory or None

				# Attribute mapping
				attr_map = {
					"metal_colour": "Metal Colour",
					"diamond_target": "Diamond Target",
					"stone_changeable": "Stone Changeable",
					"gemstone_type": "Gemstone Type",
					"chain_type": "Chain Type",
					"chain_length": "Chain Length",
					"feature": "Feature",
					"rhodium": "Rhodium",
					"enamal": "Enamal",
					"detachable": "Detachable",
					"capganthan": "Cap/Ganthan",
					"two_in_one": "Two in One",
					"product_size": "Product Size",
					"sizer_type": "Sizer Type",
					"lock_type": "Lock Type",
					"black_bead_line": "Black Bead Line",
					"charm": "Charm",
					"count_of_spiral_turns": "Count of Spiral Turns",
					"number_of_ant": "Number of Ant",
					"back_side_size": "Back Side Size",
					"space_between_mugappu": "Space between Mugappu",
					"distance_between_kadi_to_mugappu": "Distance Between Kadi To Mugappu",
					"back_belt": "Back Belt",
					"back_belt_length": "Back Belt Length",
				}

				# Clear all mapped fields if not manual override
				if not is_manual_override:
					for fieldname in attr_map.keys():
						setattr(i, fieldname, None)

				# Set values from attributes if not manually overridden
				for attr in item_doc.attributes:
					for fieldname, attrname in attr_map.items():
						if attr.attribute == attrname:
							if not is_manual_override or not getattr(i, fieldname):
								setattr(i, fieldname, attr.attribute_value)

		# Final mandatory field validation
		if  i.design_type == "New Design":
			missing = []
			if not i.category:
				missing.append("Category")
			if not i.subcategory:
				missing.append("Subcategory")
			if not i.metal_type:
				missing.append("Metal Type")
			if not i.diamond_target:
				missing.append("Diamond Target")
			if not i.setting_type:
				missing.append("Setting Type")
			if not i.metal_touch:
				missing.append("Metal Touch")
			if not i.metal_colour:
				missing.append("Metal Colour")
			if not i.metal_target:
				missing.append("Metal Target")

			if missing:
				frappe.throw(f"Row {i.idx}: Please fill the following fields for 'New Design' with Manual checked: {', '.join(missing)}")


def validate_is_mannual(self):
	if self.is_mannual:
		errors = []

		for row in self.order_details:
			missing_fields = []

			if not row.stylebio:
				missing_fields.append("'Style Bio'")
			if not row.status:
				missing_fields.append("'Status'")
			if not row.order_details_and_remarks:
				missing_fields.append("'Order Details and Remark'")

			# Enhanced: handle multiple items with same tagno, pick non-disabled one
			if row.tagno:
				matching_items = frappe.db.get_all(
					"Item",
					filters={"old_tag_no": row.tagno},
					fields=["name", "master_bom", "disabled"],
					order_by="creation desc"
				)

				selected_item = next((item for item in matching_items if not item.disabled), None)

				if selected_item:
					if not row.design_id:
						row.design_id = selected_item.name
					if not row.bom:
						row.bom = selected_item.master_bom

					if selected_item.master_bom:
						diamond_type = frappe.db.get_value(
							"BOM Diamond Detail", 
							{"parent": selected_item.master_bom}, 
							"diamond_type"
						)
						if diamond_type:
							row.diamond_type = diamond_type

			# If workflow_state == "Approved", design_type is mandatory for non-finding
			if (
				not row.is_finding_order
				and not row.design_type
				and self.workflow_state == "Approved"
			):
				missing_fields.append("'Design Type' (required in 'Creating Item & BOM')")

			if missing_fields:
				errors.append(f"Row {row.idx} is missing: {', '.join(missing_fields)}")

		if errors:
			frappe.throw("<br>".join(errors))

		# Enforce all status as 'Done' if workflow_state is Approved
		if self.workflow_state == "Approved":
			for row in self.order_details:
				if row.status != "Done":
					frappe.throw(f"Row {row.idx}: Status must be 'Done' before you approve. Please update it.")

	else:
		# is_mannual is unchecked validate design_type for non-finding orders
		missing_design_type_rows = []
		for row in self.order_details:
			if not row.is_finding_order and not row.design_type:
				missing_design_type_rows.append(
					f"Row {row.idx}: Design Type is mandatory when 'Is Finding Order' is unchecked and 'Is Mannual' is also unchecked."
				)

		if missing_design_type_rows:
			frappe.throw("<br>".join(missing_design_type_rows))



def set_data(self):
	if self.order_details:
		for i in self.order_details:
			if i.design_type in ['As Per Design Type','Mod - Old Stylebio & Tag No'] and i.design_id:
				try:
					design_id = i.design_id
					item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")
					master_bom = i.bom

					# Prepare a list to hold the item attribute names formatted as per your requirements
					all_item_attributes = []
					
					# Retrieve all item attributes for the given item subcategory
					for item_attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
						# Format the item attribute names by replacing spaces with underscores, removing '/', and converting to lower case
						formatted_attr = item_attr.item_attribute.replace(' ', '_').replace('/', '').lower()
						all_item_attributes.append(formatted_attr)
					
					# Retrieve the values for the specified attributes from the BOM
					attribute_values = frappe.db.get_value("BOM", master_bom, all_item_attributes, as_dict=1)
					
					# Dynamically set the attributes on self with the retrieved values
					for key, value in attribute_values.items():
						if str(key) == "item_category":
							key = "category"
						if str(key) == "item_subcategory":
							key = "subcategory"
						a = getattr(i, key, value)
						if a:
							continue
						else:
							setattr(i, key, value)
						# Prepare a list to hold the item attribute names formatted as per your requirements
						all_item_attributes = []
						
						# Retrieve all item attributes for the given item subcategory
						for item_attr in frappe.get_doc("Attribute Value", item_subcategory).item_attributes:
							# Format the item attribute names by replacing spaces with underscores, removing '/', and converting to lower case
							formatted_attr = item_attr.item_attribute.replace(' ', '_').replace('/', '').lower()
							
							all_item_attributes.append(formatted_attr)
						
						# Retrieve the values for the specified attributes from the BOM
						attribute_values = frappe.db.get_value("BOM", master_bom, all_item_attributes, as_dict=1)
						# Dynamically set the attributes on self with the retrieved values
						for key, value in attribute_values.items():
							if str(key) == "item_category":
								key = "category"
							if str(key) == "item_subcategory":
								key = "subcategory"
							a = getattr(i, key, value)
							if a:
								continue
							else:
								setattr(i, key, value)
				except:
					frappe.throw(f"Row {i.idx} has Issue.Check BOM first.")


def create_po(self):
	qty = 0
	po_doc = frappe.new_doc("Purchase Order")
	po_doc.supplier = self.supplier
	po_doc.transaction_date = self.delivery_date
	po_doc.company = self.company
	po_doc.branch = self.branch
	po_doc.project = self.project
	po_doc.purchase_type = 'Subcontracting'
	po_doc.schedule_date = self.delivery_date

	po_item_log = po_doc.append("items", {})
	if self.purchase_type == 'Design':
		po_item_log.item_code = "Design Expness"
	elif self.purchase_type == 'RPT':
		po_item_log.item_code = "RPT Expness"
	elif self.purchase_type == 'Model':
		total_weight = 0
		item_code = ''
		for i in self.order_details:
			if i.metal_touch == '18KT':
				item_code = "Semi Finish Goods 18KT"
			if i.metal_touch == '22KT':
				item_code = "Semi Finish Goods 22KT"
		po_item_log.item_code = item_code
	elif self.purchase_type == 'Mould':
		po_item_log.item_code = "Mould Expness"
	
	if self.purchase_type in ['Model']:
		qty_18 = 0
		qty_22 = 0
		for i in self.order_details:
			if i.metal_touch == '18KT':
				qty_18 += i.qty
			if i.metal_touch == '22KT':
				qty_22 += i.qty
		if qty_18:
			qty = qty_18
		else:
			qty = qty_22
	else:
		for i in self.order_details:
			qty+=i.qty
	
	po_item_log.qty = qty
	po_item_log.schedule_date = self.delivery_date
	po_item_log.schedule_date = self.delivery_date
	po_item_log.qty = len(self.order_details)
	po_doc.save()
	po_name = po_doc.name
	frappe.db.set_value("Purchase Order",po_name,"custom_form","Order Form")
	frappe.db.set_value("Purchase Order",po_name,"custom_form_id",self.name)
	msg = _("The following {0} is created: {1}").format(
			frappe.bold(_("Purchase Order")), "<br>" + get_link_to_form("Purchase Order", po_name)
		)
	
	frappe.msgprint(msg)
@frappe.whitelist()
def make_from_pre_order_form(source_name, target_doc=None):

    if isinstance(target_doc, str):
        target_doc = json.loads(target_doc)

    target_doc = frappe.new_doc("Order Form") if not target_doc else frappe.get_doc(target_doc)
    pre_order = frappe.db.get_value(
        "Pre Order Form",
        source_name,
        [
            "customer_code",
            "order_date",
            "sales_person",
            "diamond_quality",
            "branch",
            "order_type",
            "due_days",
            "po_no",
            "delivery_date"
        ],
        as_dict=True
    )

    if not pre_order:
        return target_doc

    target_doc.customer_code = pre_order.customer_code
    target_doc.order_date = pre_order.order_date
    target_doc.salesman_name = pre_order.sales_person
    target_doc.diamond_quality = pre_order.diamond_quality
    target_doc.branch = pre_order.branch
    target_doc.order_type = pre_order.order_type
    target_doc.due_days = pre_order.due_days
    target_doc.po_no = pre_order.po_no
    target_doc.delivery_date = pre_order.delivery_date
    target_doc.pre_order_form = source_name

    customer_order_form = frappe.db.get_all(
        "Pre Order Form Details",
        filters={"parent": source_name, "status": "Done"},
        fields="*"
    )
    for st in frappe.db.get_values("Service Type 2", {"parent": source_name}, "service_type1"):
        target_doc.append("service_type", {"service_type1": st[0]})

    for tr in frappe.db.get_values("Territory Multi Select", {"parent": source_name}, "territory"):
        target_doc.append("parcel_place", {"territory": tr[0]})

    design_ids = {i.item_variant for i in customer_order_form if i.item_variant}
    item_subcategories = set(
        frappe.db.get_values("Item", {"name": ["in", list(design_ids)]}, "item_subcategory", as_dict=False)
    )
    attribute_rows = frappe.db.get_all(
        "Attribute Value Item Attribute",
        filters={"parent": ["in", list(item_subcategories)]},
        fields=["parent", "item_attribute"]
    )

    attribute_map = {}
    for row in attribute_rows:
        formatted = row.item_attribute.replace(" ", "_").replace("/", "").lower()
        attribute_map.setdefault(row.parent, []).append(formatted)

    variant_attributes = frappe.db.get_all(
        "Item Variant Attribute",
        filters={"parent": ["in", list(design_ids)]},
        fields=["parent", "attribute", "attribute_value"]
    )

    variant_attr_map = {}
    for v in variant_attributes:
        key = v.attribute.replace(" ", "_").replace("/", "").lower()
        variant_attr_map.setdefault(v.parent, {})[key] = v.attribute_value

    bom_names = {i.bom for i in customer_order_form if i.bom}
    bom_data = frappe.db.get_all(
        "BOM",
        filters={"name": ["in", list(bom_names)]},
        fields="*"
    )
    bom_map = {b.name: b for b in bom_data}
    for i in customer_order_form:

        design_id = i.item_variant
        item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")
        master_bom = i.bom

        extra_fields = {}

        if item_subcategory and master_bom:
            allowed_attrs = attribute_map.get(item_subcategory, [])
            variant_vals = variant_attr_map.get(design_id, {})
            bom_vals = bom_map.get(master_bom, {})

            for attr in allowed_attrs:
                value = variant_vals.get(attr) or getattr(bom_vals, attr, None)
                if value:
                    fieldname = (
                        "category" if attr == "item_category"
                        else "subcategory" if attr == "item_subcategory"
                        else attr
                    )
                    extra_fields[fieldname] = value

        target_doc.append("order_details", {
            "design_by": i.design_by,
            "design_type": i.design_type,
            "order_type": i.order_type,
            "delivery_date": pre_order.delivery_date,
            "diamond_quality": pre_order.diamond_quality,
            "design_id": design_id,
            "mod_reason": i.mod_reason,
            "bom": i.bom,
            "category": i.new_category,
            "subcategory": i.new_sub_category,
            "metal_target": i.gold_target,
            "diamond_target": i.diamond_target,
            "setting_type": i.bom_setting_type,
            "pre_order_form_details": i.name,
            "diamond_type": "Natural",
            "jewelex_batch_no": i.bulk_order_no,
            "design_image_1": i.design_image,
            **({"metal_touch": i.metal_touch} if i.design_type == "New Design" else {}),
            **({"metal_colour": i.metal_color} if i.design_type == "New Design" else {}),
            **extra_fields
        })

    return target_doc


@frappe.whitelist()
def gc_export_to_excel(order_form, doc):
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)  
	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	
	file_name = f"GC_Format_{order_date_str}.xlsx" 

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'GC Format'

	# Define headers
	headers = [
		'Code on Tag','Product Category',
		'Product Wt','CFA','Brand',
		'KT','Stone size',
		'Stone Code',
		'Stone Qty','Check stock code Duplicated',
		'Brief CATPB',
		'Remarks'
		]
	
	sheet.append(headers)

	# Store all rows in a list before writing to the sheet
	rows_data = []
	
	for row in doc.get('order_details', []):
		if row.get('design_id'):
			bom_list = ''
			if row.get('tag_no'): 
				bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['*'])
			else:  
				bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': row['bom']}, fields=['*'])

			if bom_list:
				bom_diamond = frappe.db.get_all("BOM Diamond Detail", 
								filters={'parent': bom_list[0].get('name')}, fields=['*'])
				max_rows = len(bom_diamond) or 1

				for i in range(max_rows):
					diamond = bom_diamond[i] if i < len(bom_diamond) else {}
					
					row_data = [
						row.get('design_id', '') if i == 0 else "",
						row.get('category', '') if i == 0 else "", 
						f"{float(bom_list[0].get('gross_weight', 0)):0.3f}" if i == 0 else "",
						'',
						order_form_doc.customer_name if i == 0 else "",
						row.get('metal_touch', '') if i == 0 else "",
						f"{float(diamond.get('size_in_mm', 0)):0.2f} MM",
						'',
						f"{float(diamond.get('pcs', 0)):0.2f}",
						'',
						'',
						''
					]
					rows_data.append(row_data)

	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("GC Sheet Can Not Download")

	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)
	
	return file_doc.file_url

@frappe.whitelist()
def creation_export_to_excel(order_form, doc):
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)  
	
	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Code_Creation_File_{order_date_str}.xlsx" 

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Code Creation File'

	headers = [
		"S.No","Date","Collection Name","Theme Code","Designer","Karat","Complexity",
		"CFA","Vendor Name","Vendor Ref Code","Category","Group","Individual wt",
		"Total Wt","Catpb","Length","Size","Cart","Findings","Stone Quality",
		"Shape","Metal Color","UOM","Gender","Remarks","Stone Combination"
	]

	sheet.append(headers) 
	
	rows_data = []

	for row in doc.get('order_details', []):
		if row['design_id']:
			finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
			
			finish_bom = ''
			if len(finish_bom_list) > 1:
				order = frappe.db.get_value("Order", 
					{'cad_order_form': order_form, 'item': row['design_id']},'name')	
				pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
				snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
				fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
				finish_bom = fg_bom		
			else:
				for fg in finish_bom_list:
					finish_bom = fg.get('name')
			
			final_bom = ''
			if finish_bom:
				final_bom = finish_bom
			else:
				final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
					
			if final_bom:
				item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
					
				order_date = frappe.utils.formatdate(order_form_doc.order_date, "dd.MM.yyyy")
				row_data = [
					row.get('idx', ''),
					order_date,
					row.get('collection_name', '') ,
					'',
					'',
					row.get('metal_touch', ''),
					row.get('mfg_complexity_code', ''),
					'',
					order_form_doc.company,
					row.get('design_id', ''),
					row.get('category', ''),
					'',
					'', 
					'', 
					'', 
					'',
					'', 
					'',
					'', 
					row.get('diamond_quality', ''),
					'',
					row.get('metal_colour', ''),
					row.get('uomset_of', ''),
					row.get('gender', ''),
					'',
					'',
					
				]
				rows_data.append(row_data) 

	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Code creation Sheet Can Not Download , Check all details..")
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)
	
	return file_doc.file_url

@frappe.whitelist()
def proto_export_to_excel(order_form, doc):

	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Proto_Sheet_{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Proto Sheet'
	
	# Store all rows in a list before writing to the sheet
	rows_data = []
	if 'Caratlane' in order_form_doc.customer_name:
		# Define headers
		headers = [
			"Caratlane SKU Code", "Item Code", "Vendor Style Code", "Images",
			"Gold Kt", "Gold Colour", "Product Type", "Product Size", "Stone Type",
			"Diamond Sieve Size/Col Stone", "Diamond Shape", "Diamond Sieve Size(mm Size)",
			"Quantity", "Individual Stone Wt", "Total Stone Wt", "Setting Type", "Type",
			"Stone Quality", "Stone Colour", "Cut", "Rate PCT", "Value", "Gross Weight",
			"Metal Colour", "Metal Karat", "Quantity", "Gold Weight", "Finding Name",
			"Finding Quantity", "Finding Weight", "Finding Colour", "Finding Karat","Finding Type", 
			"Net Weight(Min)", "Net Weight(Avg)", "Net Weight(Max)",
			"Diamond Weight(Min)", "Diamond Weight(Avg)", "Diamond Weight(Max)",
			"Finishing Information", "Shipping Days", "Metal Rate", "Total Dia",
			"Cent per gm", "Labor", "Per Pc Labor", "Wastage", "Total", "Total Price", "Technique"
		]
		sheet.append(headers)

		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Template'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
							
				if finish_bom:
					item_image = frappe.db.get_value("Item", {'name': row["design_id"]}, ['image'])
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': finish_bom}, fields=['*'])
					bom_metal = frappe.db.get_all("BOM Metal Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_diamond = frappe.db.get_all("BOM Diamond Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_finding = frappe.db.get_all("BOM Finding Detail", filters={'parent': finish_bom}, fields=['*'])
					bom_gems = frappe.db.get_all("BOM Gemstone Detail", filters={'parent': finish_bom}, fields=['*'])
									
					# Get the maximum number of rows needed for this item
					max_rows = max(len(bom_diamond), len(bom_finding), len(bom_metal), len(bom_gems)) or 1

					for i in range(max_rows):
						diamond = bom_diamond[i] if i < len(bom_diamond) else {}
						finding = bom_finding[i] if i < len(bom_finding) else {}
						metal = bom_metal[i] if i < len(bom_metal) else {}
						gemstone = bom_gems[i] if i < len(bom_gems) else {}
						
						diamond_tolerance = set_tolerance(diamond.get('quantity', 0), order_form_doc.customer_code)
						
						row_data = [
							"",  # Caratlane SKU Code
							row.get('design_id', '') if i == 0 else "",  # Item Code (only first row)
							"",  # Vendor Style Code
							item_image if i == 0 else "",  # Images (only first row)
							metal.get('metal_touch', '') if i == 0 else "",  # Gold Kt (only first row)
							metal.get('metal_colour', '') if i == 0 else "",  # Gold Colour (only first row)
							row.get('category', '') if i == 0 else "",  # Product Type (only first row)
							row.get("product_size", "") if i == 0 else "",  # Product Size (only first row)
							diamond.get('diamond_type', ''),  # Stone Type
							diamond.get('sieve_size_range', ''),  # Diamond Sieve Size/Col Stone
							diamond.get('stone_shape', ''),  # Diamond Shape
							diamond.get('size_in_mm', ''),  # Diamond Sieve Size(mm Size)
							diamond.get('pcs', ''),  # Quantity
							diamond.get('weight_per_pcs', ''),  # Individual Stone Wt
							"",  # Total Stone Wt (not available in the given structure)
							diamond.get('sub_setting_type', ''),  # Setting Type
							"",  # Type
							diamond.get('quality', ''),  # Stone Quality
							diamond.get('sieve_size_color', ''),  # Stone Colour
							"", "", "",  # Cut, Rate PCT, Value
							item_bom[0].get('gross_weight', '') if i == 0 else "",  # Gross Weight (only first row)
							metal.get('metal_colour', ''),  # Metal Colour
							metal.get('metal_touch', ''),  # Metal Karat
							metal.get('quantity', ''),  # Quantity
							metal.get('actual_quantity', ''),  # Gold Weight
							finding.get('finding_category', ''),  # Finding Name
							finding.get('qty', ''),  # Finding Quantity
							finding.get('quantity', ''),  # Finding Weight
							finding.get('metal_colour', ''),  # Finding Colour
							finding.get('metal_touch', ''),  # Finding Karat
							"", # Finding Type
							"", "", # Net Weights
							item_bom[0].get('metal_and_finding_weight', '') if i == 0 else "",
							diamond_tolerance.get('min_diamond', ''),  # Diamond Weight (Min)
							diamond_tolerance.get('diamond_weight', ''),  # Diamond Weight (Avg)
							diamond_tolerance.get('max_diamond', ''),  # Diamond Weight (Max)
							"", "", 
							metal.get('rate'), #metal rate 
							"", "", "", "", "", "", "", "", ""  # Remaining empty fields
						]
						rows_data.append(row_data)

	
	elif 'Reliance' in order_form_doc.customer_name:
		# Define headers
		headers = [
			"Sr. NO.","Collection Name", "Vendor Name", "Vendor Design Code", "Proto Image", "Article", 
			"Metal Color", "Purity", "Stone Clarity", "Approx Net Wt (gms)", "Approx Dia Wt (cts)", 
			"Approx Color Stone Wt (cts)", "Size", "Findings", "Design Approved By", "Catrgory Approved By", 
			"Sourcing Approved By", "NPD Approved By", "QA Approved By", "QA Remarks", "Remark"
		]
		sheet.append(headers)

		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
												
				if final_bom:
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])

					realiance_quality = frappe.db.get_value("Customer Prolif Detail", 
						{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
						['customer_prolif']
						) 
					realiance_quality if realiance_quality else ''

					codes = frappe.db.get_all(
						"Reliance Size Master", 
						filters={
							'customer': order_form_doc.customer_code,
							'item_category': row.get('category')
						},
						or_filters=[
							['product_size', 'like', f"{row.get('product_size')}%"]
						],
						fields=['code','product_size'],
					)
					order_size = float(row.get('product_size'))
					
					code_categories = frappe.db.get_value(
						"Customer Category Detail",
						{
							'parent': order_form_doc.customer_code,
							'gk_category': row.get('category') ,
							'gk_sub_category': row.get('subcategory') 
					  	},
						['customer_category','customer_subcategory','code_category','article'],
						as_dict=True
					)

					row_data = [
						row.get('idx') ,
						row.get('collection_name', '') ,
						"GK", #order_form_doc.company,
						row.get('design_id', '') ,  
						"",
						code_categories['article'], # row.get('category', ''),
						f"{row.get('metal_colour', '')} {row.get('metal_type', '')}",
						row.get('metal_touch', ''),
						realiance_quality, # row.get('diamond_quality', ''),
						item_bom[0].get('metal_and_finding_weight', '') ,
						item_bom[0].get('diamond_weight', '') , 
						"",
						"", # row.get('product_size', ''),
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",						
					]
					rows_data.append(row_data)

	elif 'Novel' in order_form_doc.customer_name:
		# Define headers
		headers = [
			"SR. NO.","Design Selecion Date","Collection Name","Vendor Name","Order Type","Image","Theme Code","Vendor/ Designer Ref Code","Set Code",
			"Product Group","Product SubGroup","Product Category","Sub Category","Category, Sub-Category Code","Size","Size (UOM)","KT","Metal Color",
			"Diamond Quality","Stone Proliferation","Qty","UOM","Findings","Proto Remarks in PO","Metal Purity","Gross Wt.","Gold Weight","Diamond Carat Weight",
			"Polki Wt.","Other Stone Weight","Polki Quality","Gender","Design Source/Route","TOTAL LABOUR AMOUNT","DIAMOND HANDLING AMOUNT","TOTAL DIAMOND AMOUNT",
			"COLORSTONE HANDLING AMOUNT","COLORSTONE AMOUNT","GOLD AMOUNT","LOSS AMOUNT","ADDITIONAL CHARGES","TOTAL VALUE","Design Complexity","Need state",
			"Primary Design language","Name of the Design Motif","Modularity Flag","Modularity description","Finish Type","Colour Stone Name","Colour Stone Type",
			"Colorstone Color Family","Enamel Color Family","Bangle"
			]

		sheet.append(headers)
		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
								
				if final_bom:
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
					order_date_fmt = frappe.utils.formatdate(order_form_doc.order_date, "dd-MM-yyyy")
					
					novel_quality = frappe.db.get_value("Customer Prolif Detail", 
						{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
						['customer_prolif']
						) 
					novel_quality if novel_quality else ''
					
					product_size = row.get('product_size')
					order_size = float(product_size)

					code_entry = frappe.db.get_value(
						"Novel Size Master",
						{
							'customer': order_form_doc.customer_code,
							'item_category': row.get('category'),
							'product_size_in': product_size
						},
						['code', 'product_size'],
						as_dict=True
					)

					code_size = code_entry['code'] if code_entry else order_size

					metal_purity = float(item_bom[0].get('metal_purity', 0))
					converted_purity = round(metal_purity / 100, 2)

					code_categories = frappe.db.get_value(
						"Customer Category Detail",
						{
							'parent': order_form_doc.customer_code,
							'gk_category': row.get('category') ,
							'gk_sub_category': row.get('subcategory') 
					  	},
						['customer_category','customer_subcategory','code_category'],
						as_dict=True
					)
					
					row_data = [
						row.get('idx') ,
						order_date_fmt,
						row.get('collection_name', '') ,
						order_form_doc.company,
						f"{order_form_doc.flow_type} Order",
						"",
						"",
						row.get('design_id', ''),
						row.get('category', ''),
						"Studded",
						"Studded-DIS",
						code_categories['customer_category'], #category
						code_categories['customer_subcategory'], #subcategory
						code_categories['code_category'], #code
						code_size,
						"",
						row.get('metal_touch', ''),
						row.get('metal_colour', ''),
						novel_quality,
						"",
						row.get('qty', ''),
						row.get('uomset_of', ''),
						"", #finding
						"",
						converted_purity , #metal purity
						item_bom[0].get('gross_weight', '') , #gross wt
						item_bom[0].get('metal_and_finding_weight', 'metal_weight') , #gold wt
						item_bom[0].get('diamond_weight', '') , #diam wt
						"",
						"",
						"",
						row.get('gender', ''),
						"",
						"", #labour amount
						"", #diam handling amt
						"", #diam amt
						"", #colorstone handling amt
						"", #colorstone amt
						"", #gold amt
						"", #loss amt
						"", #additional charge
						"", #total value
						row.get('mfg_complexity_code', ''),
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						"",
						]
					rows_data.append(row_data)

	
	# Write all rows to the Excel sheet at once
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Proto Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url

@frappe.whitelist()
def get_variant_format(order_form, doc): 
	
	order_form_doc = frappe.get_doc('Order Form', order_form)
	doc = json.loads(doc)

	order_date_str = getdate(order_form_doc.order_date).strftime("%Y-%m-%d")
	file_name = f"Variant_Format_{order_date_str}.xlsx"

	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'Variant Format'

	rows_data = []

	if 'Reliance' in order_form_doc.customer_name:
		headers = [
			"Vendor Code","Article","Vendor design code", "Purity","Set of", "Metal Color", 
			"Dia quality", "Variant Size","Net Wt","Dia pcs", "Dia Wt",
			"Color Stone pcs", "Color Stone Wt", "Gross Wt", "Remark"
		]
		sheet.append(headers)

		# Loop through order details
		for row in doc.get('order_details', []):
			if row['design_id']:
				finish_bom_list = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'bom_type': 'Finish Goods'}, fields=['name'])
				
				finish_bom = ''
				if len(finish_bom_list) > 1:
					order = frappe.db.get_value("Order", 
						{'cad_order_form': order_form, 'item': row['design_id']},'name')	
					pmo = frappe.db.get_value("Parent Manufacturing Order",{'order_form_id': order}, 'name')
					snc = frappe.db.get_value("Serial Number Creator",{'parent_manufacturing_order': pmo}, 'name')
					fg_bom = frappe.db.get_value("BOM", {'custom_serial_number_creator': snc, 'item': row['design_id'], 'bom_type': 'Finish Goods'}, 'name')
					finish_bom = fg_bom		
				else:
					for fg in finish_bom_list:
						finish_bom = fg.get('name')
				
				final_bom = ''
				if finish_bom:
					final_bom = finish_bom
				else:
					final_bom = frappe.db.get_value("Item", {'name': row["design_id"],}, ['master_bom'])
												
				if final_bom:
					item_bom = frappe.db.get_list("BOM", filters={'item': row["design_id"], 'name': final_bom}, fields=['*'])
					
					realiance_quality = frappe.db.get_value("Customer Prolif Detail", 
						{'parent': order_form_doc.customer_code, 'gk_d': row.get('diamond_quality')  },
						['customer_prolif']
						) 
					realiance_quality if realiance_quality else ''

					code_categories = frappe.db.get_value(
						"Customer Category Detail",
						{
							'parent': order_form_doc.customer_code,
							'gk_category': row.get('category') ,
							'gk_sub_category': row.get('subcategory') 
					  	},
						['customer_category','customer_subcategory','code_category','article'],
						as_dict=True
					)

					row_data = [
						"",
						code_categories['code_category'],
						row.get('design_id', '') ,  
						row.get('metal_touch', ''),
						"",
						f"{row.get('metal_colour', '')} {row.get('metal_type', '')}",
						realiance_quality,
						"", 
						item_bom[0].get('metal_and_finding_weight', '') ,
						item_bom[0].get('total_diamond_pcs', '') ,
						item_bom[0].get('diamond_weight', '') , 
						"",
						"",
						item_bom[0].get('gross_weight', '') , 
						"",

					]
					rows_data.append(row_data)
				# Write all rows to the Excel sheet at once
	
	if rows_data:
		for row in rows_data:
			sheet.append(row)
	else:
		frappe.throw("Proto Sheet Can Not Download")

	# Save the workbook to a BytesIO stream
	output = BytesIO()
	workbook.save(output)
	output.seek(0)

	file_doc = save_file(
		file_name,
		output.getvalue(),
		order_form_doc.doctype,
		order_form_doc.name,
		is_private=0
	)

	return file_doc.file_url

def set_tolerance(diamond_weight, customer):
	data_json = {}
	if diamond_weight:
		tolerance_data = frappe.db.get_all('Diamond Tolerance Table',
			filters={'weight_type': 'Weight wise', 'parent': customer}, 
			fields=['from_diamond', 'to_diamond', 'plus_percent', 'minus_percent'])

		for row in tolerance_data:
			if row['from_diamond'] <= diamond_weight <= row['to_diamond']:
				plus_percent = row['plus_percent']
				minus_percent = row['minus_percent']

				max_diamond_weight = diamond_weight + plus_percent
				min_diamond_weight = diamond_weight - minus_percent
				
				data_json['diamond_weight'] = round(diamond_weight, 3)
				data_json['max_diamond'] = round(max_diamond_weight, 3)
				data_json['min_diamond'] = round(min_diamond_weight, 3)
				
	return data_json


@frappe.whitelist()
def get_bom_details(design_id, doc):
	
	doc = json.loads(doc)

	if doc.get("is_finding_order"):
		master_bom = frappe.db.get_value(
		"BOM",
		{"bom_type": "Template", "item": design_id},
		"name",
		order_by="creation DESC"
		)
		frappe.throw(f"{master_bom}//{doc['is_finding_order']}")

	item_subcategory = frappe.db.get_value("Item", design_id, "item_subcategory")

	fg_bom = frappe.db.get_value(
		"BOM",
		{"bom_type": "Finished Goods", "item": design_id},
		"name",
		order_by="creation DESC"
	)
	master_bom = fg_bom or frappe.db.get_value("Item", design_id, "master_bom")

	if not master_bom:
		frappe.throw(
		f"Master BOM for Item <b>{get_link_to_form('Item', design_id)}</b> is not set"
		)

	item_attributes = frappe.db.get_all(
		"Item Attribute Detail",
		filters={"parent": item_subcategory},
		fields=["item_attribute"]
	)

	attribute_pairs = []
	attribute_keys = []

	for row in item_attributes:
		formatted = (
		row.item_attribute
		.replace(" ", "_")
		.replace("/", "")
		.lower()
		)
		attribute_pairs.append((row.item_attribute, formatted))
		attribute_keys.append(formatted)

	# Variant attributes
	variant_attributes = frappe.db.get_all(
		"Item Variant Attribute",
		filters={"parent": design_id},
		fields=["attribute", "attribute_value"]
	)

	variant_map = {
		row.attribute.replace(" ", "_").replace("/", "").lower(): row.attribute_value
		for row in variant_attributes
	}

	# BOM fallback values
	bom_values = frappe.db.get_value(
		"BOM",
		master_bom,
		attribute_keys,
		as_dict=True
	) or {}

	with_value = {}
	for original, key in attribute_pairs:
		with_value[key] = variant_map.get(key) or bom_values.get(key)

	with_value["master_bom"] = master_bom
	return with_value
